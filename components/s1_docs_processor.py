"""
SentinelOne Documentation Processor for RAG Knowledge Base
Comprehensive processing of S1 query language, OCSF schemas, APIs, and field mappings
"""
import logging
import json
import re
import csv
from pathlib import Path
from typing import Dict, List, Any, Optional, Tuple
from datetime import datetime
import PyPDF2
import yaml

logger = logging.getLogger(__name__)


class S1DocsManager:
    """
    Comprehensive SentinelOne documentation processor
    Handles PDFs, JSON schemas, CSV data dictionaries, and YAML API specs
    """

    def __init__(self, docs_directory: str, rag_knowledge_base=None):
        """
        Initialize S1 documentation manager

        Args:
            docs_directory: Path to S1 documentation directory
            rag_knowledge_base: Optional RAG knowledge base for ingestion
        """
        self.docs_dir = Path(docs_directory)
        self.rag = rag_knowledge_base
        self.processing_stats = {
            "files_processed": 0,
            "chunks_created": 0,
            "chunks_ingested": 0,
            "field_mappings_extracted": 0,
            "query_patterns_extracted": 0,
            "api_endpoints_extracted": 0,
            "errors": []
        }

    async def process_and_ingest(self) -> Dict[str, Any]:
        """
        Process all S1 documentation and ingest into RAG

        Returns:
            Processing statistics
        """
        logger.info("[SEARCH] Starting S1 documentation processing...")

        all_chunks = []

        # Process PDFs
        pdf_chunks = await self._process_all_pdfs()
        all_chunks.extend(pdf_chunks)

        # Process GraphQL schema
        graphql_chunks = await self._process_graphql_schema()
        all_chunks.extend(graphql_chunks)

        # Process vulnerabilities CSV
        csv_chunks = await self._process_vulnerability_csv()
        all_chunks.extend(csv_chunks)

        # Process Swagger/OpenAPI specs
        swagger_chunks = await self._process_swagger_specs()
        all_chunks.extend(swagger_chunks)

        # Process Excel OCSF mapping
        excel_chunks = await self._process_ocsf_excel()
        all_chunks.extend(excel_chunks)

        self.processing_stats["chunks_created"] = len(all_chunks)

        # Ingest into RAG
        if self.rag and self.rag.enabled:
            logger.info(f"[INBOX] Ingesting {len(all_chunks)} chunks into RAG...")
            for chunk in all_chunks:
                try:
                    self.rag.add_document(
                        content=chunk["content"],
                        title=chunk["title"],
                        source=chunk["source"],
                        doc_type=chunk["doc_type"],
                        metadata=chunk.get("metadata", {})
                    )
                    self.processing_stats["chunks_ingested"] += 1
                except Exception as e:
                    error_msg = f"Failed to ingest chunk: {e}"
                    logger.error(error_msg)
                    self.processing_stats["errors"].append(error_msg)

        return {
            "summary": self.processing_stats,
            "chunks": all_chunks[:10]  # Sample of chunks
        }

    async def _process_all_pdfs(self) -> List[Dict[str, Any]]:
        """Process all PDF documents"""
        logger.info("[NOTE] Processing PDF documents...")

        chunks = []
        pdf_files = list(self.docs_dir.glob("*.pdf"))

        for pdf_path in pdf_files:
            try:
                logger.info(f"  Processing: {pdf_path.name}")
                pdf_chunks = await self._process_pdf(pdf_path)
                chunks.extend(pdf_chunks)
                self.processing_stats["files_processed"] += 1
            except Exception as e:
                error_msg = f"Failed to process PDF {pdf_path.name}: {e}"
                logger.error(error_msg)
                self.processing_stats["errors"].append(error_msg)

        logger.info(f"[OK] Processed {len(pdf_files)} PDFs → {len(chunks)} chunks")
        return chunks

    async def _process_pdf(self, pdf_path: Path) -> List[Dict[str, Any]]:
        """
        Process individual PDF with intelligent chunking

        Args:
            pdf_path: Path to PDF file

        Returns:
            List of document chunks with metadata
        """
        chunks = []

        try:
            with open(pdf_path, 'rb') as file:
                pdf_reader = PyPDF2.PdfReader(file)
                full_text = ""

                # Extract all text
                for page_num, page in enumerate(pdf_reader.pages):
                    page_text = page.extract_text()
                    full_text += f"\n\n--- Page {page_num + 1} ---\n\n{page_text}"

                # Detect document type and process accordingly
                doc_type = self._detect_s1_doc_type(pdf_path.name, full_text)

                if "query" in doc_type.lower():
                    chunks.extend(self._extract_query_patterns(full_text, pdf_path.name))
                elif "operator" in doc_type.lower():
                    chunks.extend(self._extract_operators(full_text, pdf_path.name))
                elif "ocsf" in doc_type.lower():
                    chunks.extend(self._extract_ocsf_mappings(full_text, pdf_path.name))
                elif "field" in doc_type.lower():
                    chunks.extend(self._extract_field_definitions(full_text, pdf_path.name))
                else:
                    # Generic chunking
                    chunks.extend(self._chunk_document(full_text, pdf_path.name, doc_type))

        except Exception as e:
            logger.error(f"Error processing PDF {pdf_path.name}: {e}")

        return chunks

    def _detect_s1_doc_type(self, filename: str, content: str) -> str:
        """Detect type of S1 documentation"""
        filename_lower = filename.lower()

        if "query" in filename_lower:
            if "language" in filename_lower or "overview" in filename_lower:
                return "s1_query_language"
            elif "field" in filename_lower:
                return "s1_query_fields"
            else:
                return "s1_querying"
        elif "operator" in filename_lower:
            return "s1_operators"
        elif "ocsf" in filename_lower:
            if "event" in filename_lower or "vulnerabilit" in filename_lower:
                return "ocsf_events"
            elif "map" in filename_lower:
                return "ocsf_mapping"
            elif "schema" in filename_lower:
                return "ocsf_schema"
            else:
                return "ocsf_general"
        elif "xdr" in filename_lower or "ingestion" in filename_lower:
            return "xdr_ingestion"
        else:
            return "s1_documentation"

    def _extract_query_patterns(self, text: str, source: str) -> List[Dict[str, Any]]:
        """Extract SentinelOne query patterns and syntax"""
        chunks = []

        # Pattern 1: Field syntax examples (e.g., field = value, field IN (values))
        field_patterns = re.finditer(
            r'([a-zA-Z_][a-zA-Z0-9_.]*)\s+(=|!=|>|<|>=|<=|IN|CONTAINS|STARTSWITH|ENDSWITH|LIKE)\s+([^;\n]+)',
            text,
            re.IGNORECASE
        )

        for match in field_patterns:
            field, operator, value = match.groups()
            chunks.append({
                "content": f"S1 Query Pattern: {field} {operator} {value}\n\nField: {field}\nOperator: {operator}\nExample Value: {value.strip()}",
                "title": f"S1 Query: {field} {operator}",
                "source": source,
                "doc_type": "s1_query_pattern",
                "metadata": {
                    "field": field,
                    "operator": operator,
                    "value": value.strip(),
                    "query_type": "field_query"
                }
            })
            self.processing_stats["query_patterns_extracted"] += 1

        # Pattern 2: Logical operators (AND, OR, NOT)
        logical_sections = re.finditer(
            r'(AND|OR|NOT)\s+(?:operator)?[:\s]+([^.\n]{10,200})',
            text,
            re.IGNORECASE
        )

        for match in logical_sections:
            operator, description = match.groups()
            chunks.append({
                "content": f"S1 Logical Operator: {operator}\n\n{description.strip()}",
                "title": f"S1 Logic: {operator}",
                "source": source,
                "doc_type": "s1_logical_operator",
                "metadata": {
                    "operator": operator,
                    "category": "logical"
                }
            })

        # Pattern 3: Function calls (e.g., COUNT(), MAX(), etc.)
        function_patterns = re.finditer(
            r'([A-Z_]+)\(([^)]*)\)',
            text
        )

        for match in function_patterns:
            func_name, func_args = match.groups()
            if func_name in ['COUNT', 'MAX', 'MIN', 'AVG', 'SUM', 'DISTINCT']:
                chunks.append({
                    "content": f"S1 Function: {func_name}({func_args})\n\nAggregation function for query results.",
                    "title": f"S1 Function: {func_name}",
                    "source": source,
                    "doc_type": "s1_function",
                    "metadata": {
                        "function": func_name,
                        "arguments": func_args
                    }
                })

        return chunks

    def _extract_operators(self, text: str, source: str) -> List[Dict[str, Any]]:
        """Extract operator definitions and comparisons"""
        chunks = []

        # Find operator comparison tables
        table_pattern = r'(\w+)\s+\|\s+([^|\n]+)\s+\|\s+([^|\n]+)'
        matches = re.finditer(table_pattern, text)

        for match in matches:
            operator, event_search, deep_visibility = match.groups()
            chunks.append({
                "content": f"Operator: {operator}\nEvent Search: {event_search.strip()}\nDeep Visibility: {deep_visibility.strip()}",
                "title": f"Operator: {operator}",
                "source": source,
                "doc_type": "s1_operator_definition",
                "metadata": {
                    "operator": operator,
                    "event_search": event_search.strip(),
                    "deep_visibility": deep_visibility.strip()
                }
            })

        return chunks

    def _extract_ocsf_mappings(self, text: str, source: str) -> List[Dict[str, Any]]:
        """Extract OCSF field mappings"""
        chunks = []

        # Extract field mapping patterns
        # Format: OCSF field -> S1 field
        mapping_pattern = r'([a-z_]+(?:\.[a-z_]+)*)\s*(?:->|→|maps to|=)\s*([a-z_]+(?:\.[a-z_]+)*)'
        matches = re.finditer(mapping_pattern, text, re.IGNORECASE)

        for match in matches:
            ocsf_field, s1_field = match.groups()
            chunks.append({
                "content": f"OCSF Mapping: {ocsf_field} → {s1_field}\n\nOCSF standardized field maps to SentinelOne internal field.",
                "title": f"OCSF: {ocsf_field}",
                "source": source,
                "doc_type": "ocsf_field_mapping",
                "metadata": {
                    "ocsf_field": ocsf_field,
                    "s1_field": s1_field,
                    "mapping_type": "field"
                }
            })
            self.processing_stats["field_mappings_extracted"] += 1

        return chunks

    def _extract_field_definitions(self, text: str, source: str) -> List[Dict[str, Any]]:
        """Extract field definitions and metadata"""
        chunks = []

        # Pattern: field_name (type): description
        field_def_pattern = r'([a-z_][a-z0-9_.]*)\s*\(([^)]+)\)\s*[:-]\s*([^.\n]{10,300})'
        matches = re.finditer(field_def_pattern, text, re.IGNORECASE)

        for match in matches:
            field_name, field_type, description = match.groups()
            chunks.append({
                "content": f"Field: {field_name}\nType: {field_type}\nDescription: {description.strip()}",
                "title": f"Field: {field_name}",
                "source": source,
                "doc_type": "s1_field_definition",
                "metadata": {
                    "field_name": field_name,
                    "field_type": field_type.strip(),
                    "description": description.strip()
                }
            })
            self.processing_stats["field_mappings_extracted"] += 1

        return chunks

    def _chunk_document(self, text: str, source: str, doc_type: str) -> List[Dict[str, Any]]:
        """Generic document chunking with overlap"""
        chunks = []
        chunk_size = 1000  # characters
        overlap = 200

        # Split by sections if possible
        sections = re.split(r'\n\s*(?:Chapter|Section|\d+\.)\s+', text)

        for i, section in enumerate(sections):
            if len(section.strip()) < 50:
                continue

            # Further chunk large sections
            for j in range(0, len(section), chunk_size - overlap):
                chunk_text = section[j:j + chunk_size]
                if len(chunk_text.strip()) < 50:
                    continue

                chunks.append({
                    "content": chunk_text.strip(),
                    "title": f"{source} - Section {i+1} Part {j // (chunk_size - overlap) + 1}",
                    "source": source,
                    "doc_type": doc_type,
                    "metadata": {
                        "section": i + 1,
                        "chunk_index": j // (chunk_size - overlap)
                    }
                })

        return chunks

    async def _process_graphql_schema(self) -> List[Dict[str, Any]]:
        """Process GraphQL schema file (Rules and policies Schema.txt)"""
        logger.info("[SEARCH] Processing GraphQL schema...")

        chunks = []
        schema_file = self.docs_dir / "Rules and policies Schema.txt"

        if not schema_file.exists():
            logger.warning(f"GraphQL schema not found: {schema_file}")
            return chunks

        try:
            with open(schema_file, 'r', encoding='utf-8') as f:
                schema_content = f.read()

            # Extract type definitions
            type_pattern = r'type\s+(\w+)\s*{([^}]+)}'
            for match in re.finditer(type_pattern, schema_content):
                type_name, fields = match.groups()
                chunks.append({
                    "content": f"GraphQL Type: {type_name}\n\nFields:\n{fields.strip()}",
                    "title": f"GraphQL: {type_name}",
                    "source": "Rules and policies Schema.txt",
                    "doc_type": "graphql_type",
                    "metadata": {
                        "type_name": type_name,
                        "schema_type": "graphql"
                    }
                })

            # Extract enum definitions
            enum_pattern = r'enum\s+(\w+)\s*{([^}]+)}'
            for match in re.finditer(enum_pattern, schema_content):
                enum_name, values = match.groups()
                chunks.append({
                    "content": f"GraphQL Enum: {enum_name}\n\nValues:\n{values.strip()}",
                    "title": f"Enum: {enum_name}",
                    "source": "Rules and policies Schema.txt",
                    "doc_type": "graphql_enum",
                    "metadata": {
                        "enum_name": enum_name,
                        "schema_type": "graphql"
                    }
                })

            # Extract input types
            input_pattern = r'input\s+(\w+)\s*{([^}]+)}'
            for match in re.finditer(input_pattern, schema_content):
                input_name, fields = match.groups()
                chunks.append({
                    "content": f"GraphQL Input: {input_name}\n\nFields:\n{fields.strip()}",
                    "title": f"Input: {input_name}",
                    "source": "Rules and policies Schema.txt",
                    "doc_type": "graphql_input",
                    "metadata": {
                        "input_name": input_name,
                        "schema_type": "graphql"
                    }
                })

            self.processing_stats["files_processed"] += 1
            logger.info(f"[OK] Processed GraphQL schema → {len(chunks)} chunks")

        except Exception as e:
            error_msg = f"Failed to process GraphQL schema: {e}"
            logger.error(error_msg)
            self.processing_stats["errors"].append(error_msg)

        return chunks

    async def _process_vulnerability_csv(self) -> List[Dict[str, Any]]:
        """Process vulnerabilities data dictionary CSV"""
        logger.info("[SEARCH] Processing vulnerabilities CSV...")

        chunks = []
        csv_file = self.docs_dir / "Vulnerabilities data dictionary - Sheet1.csv"

        if not csv_file.exists():
            logger.warning(f"Vulnerabilities CSV not found: {csv_file}")
            return chunks

        try:
            with open(csv_file, 'r', encoding='utf-8') as f:
                csv_reader = csv.DictReader(f)
                for row in csv_reader:
                    # Create comprehensive field documentation
                    attribute = row.get('Vulnerability attribute', '')
                    data_type = row.get('Data type', '')
                    mandatory = row.get('Mandatory', '')
                    data_source = row.get('Data source', '')
                    graphql_detail = row.get('xSPM Mapping - GraphQL detail field', '')
                    graphql_list = row.get('xSPM Mapping - GraphQL list field', '')
                    available_event_search = row.get('Available in event search', '')

                    if not attribute:
                        continue

                    content = f"""Vulnerability Field: {attribute}

Data Type: {data_type}
Mandatory: {mandatory}
Data Source: {data_source}
Available in Event Search: {available_event_search}

GraphQL Mappings:
- Detail Field: {graphql_detail}
- List Field: {graphql_list}

This field is used in SentinelOne vulnerability management and can be queried via GraphQL API."""

                    chunks.append({
                        "content": content,
                        "title": f"Vuln Field: {attribute}",
                        "source": "Vulnerabilities data dictionary - Sheet1.csv",
                        "doc_type": "vulnerability_field",
                        "metadata": {
                            "attribute": attribute,
                            "data_type": data_type,
                            "mandatory": mandatory == 'Y',
                            "data_source": data_source,
                            "graphql_detail_field": graphql_detail,
                            "graphql_list_field": graphql_list,
                            "event_search_available": available_event_search == 'Y'
                        }
                    })
                    self.processing_stats["field_mappings_extracted"] += 1

            self.processing_stats["files_processed"] += 1
            logger.info(f"[OK] Processed vulnerabilities CSV → {len(chunks)} chunks")

        except Exception as e:
            error_msg = f"Failed to process vulnerabilities CSV: {e}"
            logger.error(error_msg)
            self.processing_stats["errors"].append(error_msg)

        return chunks

    async def _process_swagger_specs(self) -> List[Dict[str, Any]]:
        """Process Swagger/OpenAPI specification files"""
        logger.info("[SEARCH] Processing Swagger/OpenAPI specs...")

        chunks = []

        # Process YAML Swagger spec
        yaml_file = self.docs_dir / "sdl-api-swagger.yml"
        if yaml_file.exists():
            try:
                with open(yaml_file, 'r', encoding='utf-8') as f:
                    swagger_spec = yaml.safe_load(f)

                # Extract API paths/endpoints
                for path, methods in swagger_spec.get('paths', {}).items():
                    for method, details in methods.items():
                        if isinstance(details, dict):
                            summary = details.get('summary', '')
                            description = details.get('description', '')

                            content = f"""SDL API Endpoint: {method.upper()} {path}

Summary: {summary}

Description:
{description}

This endpoint is part of the SentinelOne Singularity Data Lake (SDL) API."""

                            chunks.append({
                                "content": content,
                                "title": f"SDL API: {method.upper()} {path}",
                                "source": "sdl-api-swagger.yml",
                                "doc_type": "api_endpoint",
                                "metadata": {
                                    "path": path,
                                    "method": method.upper(),
                                    "summary": summary,
                                    "api_type": "SDL"
                                }
                            })
                            self.processing_stats["api_endpoints_extracted"] += 1

                self.processing_stats["files_processed"] += 1
                logger.info(f"[OK] Processed SDL API spec → {len(chunks)} endpoints")

            except Exception as e:
                error_msg = f"Failed to process SDL API spec: {e}"
                logger.error(error_msg)
                self.processing_stats["errors"].append(error_msg)

        # Note: JSON swagger file is too large (13.2MB), skip for now
        # Can be processed in chunks if needed

        return chunks

    async def _process_ocsf_excel(self) -> List[Dict[str, Any]]:
        """Process OCSF Schema XDR Map Excel file"""
        logger.info("[SEARCH] Processing OCSF Excel mapping...")

        chunks = []
        excel_file = self.docs_dir / "OCSF Schema XDR Map (1).xlsx"

        if not excel_file.exists():
            logger.warning(f"OCSF Excel not found: {excel_file}")
            return chunks

        try:
            import openpyxl
            wb = openpyxl.load_workbook(excel_file)

            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]

                # Extract headers
                headers = []
                for cell in sheet[1]:
                    if cell.value:
                        headers.append(str(cell.value))

                # Process rows
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    if not any(row):
                        continue

                    row_dict = dict(zip(headers, row))
                    if 'field' in str(headers[0]).lower() or 'name' in str(headers[0]).lower():
                        field_name = row[0]
                        if field_name:
                            content = f"OCSF Mapping: {field_name}\n\n"
                            for header, value in row_dict.items():
                                if value:
                                    content += f"{header}: {value}\n"

                            chunks.append({
                                "content": content,
                                "title": f"OCSF: {field_name}",
                                "source": f"OCSF Schema XDR Map - {sheet_name}",
                                "doc_type": "ocsf_excel_mapping",
                                "metadata": {
                                    "sheet": sheet_name,
                                    "field": field_name,
                                    **row_dict
                                }
                            })
                            self.processing_stats["field_mappings_extracted"] += 1

            self.processing_stats["files_processed"] += 1
            logger.info(f"[OK] Processed OCSF Excel → {len(chunks)} mappings")

        except ImportError:
            logger.warning("openpyxl not installed, skipping Excel processing")
        except Exception as e:
            error_msg = f"Failed to process OCSF Excel: {e}"
            logger.error(error_msg)
            self.processing_stats["errors"].append(error_msg)

        return chunks

    def get_document_summary(self) -> Dict[str, Any]:
        """Get summary of available S1 documentation"""
        files = []
        total_size = 0

        for file_path in self.docs_dir.iterdir():
            if file_path.is_file():
                size_kb = file_path.stat().st_size / 1024
                total_size += size_kb

                # Count lines for text files
                lines = 0
                if file_path.suffix in ['.txt', '.csv', '.yml', '.yaml']:
                    try:
                        with open(file_path, 'r', encoding='utf-8') as f:
                            lines = sum(1 for _ in f)
                    except (IOError, OSError, UnicodeDecodeError) as e:
                        logger.warning(f"Failed to count lines in {file_path.name}: {e}")
                        lines = 0

                files.append({
                    "name": file_path.name,
                    "size_kb": round(size_kb, 2),
                    "type": file_path.suffix,
                    "lines": lines
                })

        files.sort(key=lambda x: x['size_kb'], reverse=True)

        return {
            "total_files": len(files),
            "total_size_kb": round(total_size, 2),
            "files": files
        }
